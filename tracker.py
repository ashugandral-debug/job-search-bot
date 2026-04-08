"""
Excel tracker — maintains job application status spreadsheet
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

TRACKER_FILE = "Job_Applications_Tracker.xlsx"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _ensure_tracker():
    if os.path.exists(TRACKER_FILE):
        return openpyxl.load_workbook(TRACKER_FILE)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "AJATYA GANDRAL — PLANNING ENGINEER JOB APPLICATIONS TRACKER"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws["A1"].fill      = PatternFill("solid", start_color="EBF3FB")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Email: ashugandral@gmail.com  |  Role: Planning Engineer  |  Location: Dubai, UAE"
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6

    headers = ["#", "Company", "HR Email", "Role", "Date Applied",
               "Status", "Reply", "Interview (IST)", "Follow-Up", "Notes"]
    ws.row_dimensions[4].height = 26
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()

    widths = [4, 34, 30, 24, 18, 16, 14, 22, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(TRACKER_FILE)
    return wb

def update_tracker(company, hr_email, role, status,
                   reply="Pending ⏳", interview="—", followup="—", notes=""):
    wb = _ensure_tracker()
    ws = wb.active

    # Find existing row
    for row in ws.iter_rows(min_row=5):
        if row[1].value == company:
            row[5].value  = status
            row[6].value  = reply
            row[7].value  = interview
            row[8].value  = followup
            if notes:
                row[9].value = notes
            _style_status_row(row, status)
            wb.save(TRACKER_FILE)
            return

    # Add new row
    idx     = ws.max_row - 3  # account for header rows
    alt     = "EBF3FB" if idx % 2 == 0 else "FFFFFF"
    new_row = [
        idx, company, hr_email, role,
        datetime.now().strftime("%d-%b-%Y %H:%M"),
        status, reply, interview, followup, notes
    ]
    ws.row_dimensions[ws.max_row + 1].height = 22
    ws.append(new_row)
    row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row))[0]
    for cell in row:
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=alt)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    _style_status_row(row, status)
    wb.save(TRACKER_FILE)

def _style_status_row(row, status):
    sent_fill       = PatternFill("solid", start_color="C6EFCE")
    interview_fill  = PatternFill("solid", start_color="FFEB9C")
    pending_fill    = PatternFill("solid", start_color="FCE4D6")
    if "Sent" in status:
        row[5].fill = sent_fill
    elif "Interview" in status or "🎯" in status:
        row[5].fill = interview_fill
    elif "Follow" in status:
        row[5].fill = PatternFill("solid", start_color="DDEBF7")

def get_pending_followups():
    """Return jobs that need follow-up (sent 5+ days ago, no reply)"""
    from datetime import timedelta
    wb = _ensure_tracker()
    ws = wb.active
    due = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[1]:
            continue
        company, hr_email, role, date_applied, status, reply, _, followup, _ = \
            row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]
        if status and "Sent" in str(status) and str(reply) == "Pending ⏳" and str(followup) == "—":
            try:
                applied = datetime.strptime(str(date_applied).split(" ")[0], "%d-%b-%Y")
                if (datetime.now() - applied).days >= 5:
                    due.append({"company": company, "hr_email": hr_email, "role": role})
            except:
                pass
    return due
